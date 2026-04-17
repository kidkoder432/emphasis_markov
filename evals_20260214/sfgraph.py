import openpyxl
import json
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go

name = input("Name: ")

wb = openpyxl.load_workbook(f"./0212_eval_log_{name}.xlsx")
sheet = wb.sheetnames[-1]
ws = wb[sheet]

if name == "amrit":
    cells = ws["B3:E17"]
    std = ws["B24:E38"]
else:
    cells = ws["B3:E20"]
    std = ws["B27:G44"]

cells = [[i.value for i in j] for j in cells]
std = [[i.value for i in j] for j in std]

std = pd.DataFrame(std).transpose()
cells = pd.DataFrame(cells).transpose()

cells.loc[cells.shape[0]] = 0
std.loc[std.shape[0]] = 0

labels = json.load(open(f"../raga_data_{name}/{name}.json"))["notes"]
models = ["Full Model", "Emphasis + Tags", "Emphasis-Only", "Baseline", "Dataset"]

# --- IEEE SINGLE COLUMN CALIBRATION ---
COLUMN_WIDTH_INCHES = 3.5
DPI = 300  # Higher DPI for a sharp PNG export
PX_WIDTH = int(COLUMN_WIDTH_INCHES * DPI)  # 1050px
PX_HEIGHT = int(PX_WIDTH * 0.5)  # Maintains a clean aspect ratio
FONT_SIZE = 29  # Scaled for 300 DPI to look like ~9pt in LaTeX

fig = go.Figure()

# High-contrast color palette (Modified Dark2 for better white-theme visibility)
colors = [
    "#1b9e77",  # Full Model: Teal/Green
    "#d95f02",  # Emphasis + Tags: Orange
    "#7570b3",  # Emphasis-Only: Purple
    "#e7298a",  # Baseline: Pink
    "#666666",  # Dataset: Grey
]

offset_strength = 0.12
offsets = np.linspace(-offset_strength, offset_strength, len(models))

for i, model in enumerate(models):
    y = cells.iloc[i, :].to_numpy(dtype=float)
    err = std.iloc[i, :].to_numpy(dtype=float)
    color = colors[i % len(colors)]
    x = np.arange(len(labels)) + offsets[i]

    # Specific styling for the Baseline to make it stand out from the "good" models
    line_style = dict(width=3, color=color)
    if model == "Baseline":
        line_style["dash"] = "dash"  # Makes baseline distinct even in grayscale

    fig.add_scatter(
        x=x,
        y=y,
        mode="lines",
        line=line_style,
        error_y=dict(
            array=err,
            visible=True,
            thickness=1.5,
            width=4,
            color=color,
        ),
        name=model,
    )

fig.update_layout(
    width=PX_WIDTH,
    height=PX_HEIGHT,
    template="plotly_white",
    # Margins expanded slightly for the higher resolution PNG
    margin=dict(l=100, r=20, t=20, b=100),
    font=dict(
        family="Verdana",
        size=FONT_SIZE,
        color="black",
    ),
    showlegend=False,
    xaxis=dict(
        tickmode="array",
        tickvals=np.arange(len(labels)),
        ticktext=labels,
        tickfont=dict(size=FONT_SIZE - 6),
        title_text="Note",
        title_standoff=20,
        linecolor="black",
        mirror=True,  # Adds box-style border
    ),
    yaxis=dict(
        title_text="Euclidean Dist.",
        title_standoff=20,
        tickfont=dict(size=FONT_SIZE - 6),
        linecolor="black",
        mirror=True,
        gridcolor="lightgrey",  # Soft grid for readability
    ),
)

# Export as high-res PNG
fig.write_image(f"../0212_evals_{name}.png", scale=1, width=PX_WIDTH, height=PX_HEIGHT)
