from turtle import width
import openpyxl
import json
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go

name = input("Name: ")

wb = openpyxl.load_workbook(f"./1019_evals_{name}.xlsx")
sheet = wb.sheetnames[-1]
ws = wb[sheet]

if name == "amrit":
    cells = ws["B3:E17"]
    std = ws["B24:E38"]
else:
    cells = ws["B3:E20"]
    std = ws["B25:G42"]

cells = [[i.value for i in j] for j in cells]
std = [[i.value for i in j] for j in std]

std = pd.DataFrame(std).transpose()
cells = pd.DataFrame(cells).transpose()

cells.loc[cells.shape[0]] = 0
std.loc[std.shape[0]] = 0

labels = json.load(open(f"../raga_data_{name}/{name}.json"))["notes"]
models = ["Full Model", "Emphasis + Tags", "Emphasis-Only", "Baseline", "Dataset"]

fig = go.Figure()
colors = px.colors.qualitative.Dark2

# spacing factor — small shifts so they don't overlap
offset_strength = 0.12  # tweak this for how separated you want them
offsets = np.linspace(-offset_strength, offset_strength, len(models))

for i, model in enumerate(models):
    y = cells.iloc[i, :].to_numpy(dtype=float)
    err = std.iloc[i, :].to_numpy(dtype=float)
    color = colors[i % len(colors)]

    # offset x positions slightly for each model
    x = np.arange(len(labels)) + offsets[i]

    fig.add_scatter(
        x=x,
        y=y,
        mode="lines+markers",
        marker_color=color,
        line_color=color,
        error_y=dict(array=err, visible=True, color=color),
        name=model,
    )

fig.update_layout(
    xaxis=dict(tickmode="array", tickvals=np.arange(len(labels)), ticktext=labels),
    xaxis_title="Note",
    yaxis_title="Euclidean Distance",
    template="plotly_white",
    font=dict(
        family="Verdana",
        size=22,
        color="black",
    ),
)

fig.show()
fig.write_image(f"../1019_evals_{name}.svg", format="svg", width=1600, height=800)
